import openpyxl

# Criar planilha 
book = openpyxl.Workbook()

#Visualizar páginas existentes
print(book.sheetnames)

#Criar página
book.create_sheet('Frutas')

#Selecionar página
frutas_page = book['Frutas']
frutas_page.append(['Frutas', 'Quantidade', 'Preço'])

frutas_page.append(['Banada', '5', 'R$3, 89'])
frutas_page.append(['Melão', '3', 'R$5, 76'])
frutas_page.append(['Uva', '7', 'R$13, 70'])
frutas_page.append(['Maça', '2', 'R$6, 90'])
frutas_page.append(['Manga', '1', 'R$5, 90'])

#Salvar a planilha
book.save('Planilha de compras.xlsx')